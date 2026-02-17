#!/usr/bin/env python3
"""
Backfill missing editing_efficiency for PMC12008803 (paper_id=44).

Reads raw Excel files (mmc4, mmc6, mmc8) and matches DB entries by row number
from raw_source_text. Picks the best available efficiency value from multiple
timepoint/replicate columns.

Summary:
- mmc2: Already complete (3,892/3,892)
- mmc4 (SMARCB1): 1,890 missing - uses D10 average, fallback to D20, D4, etc.
- mmc6 (MLH1 exon 10): 1,670 missing - uses PEmax_D20, fallback to others
- mmc8 (MLH1 non-coding): 2,934 missing - uses PEmax_D20, fallback to others
"""

import re
import sqlite3
import openpyxl


DB_PATH = "data/pegrna_database.db"
RAW_DIR = "data/raw_papers/PMC12008803/PMC12008803"

# Efficiency column preferences per file (skip neg_control and plasmid)
MMC4_EFF_COLS = [
    # Prefer D10 replicates (mid-timepoint)
    ("D10_R1_percentage_editing", "D10_R2_percentage_editing"),
    # Fallback to D20
    ("D20_R1_percentage_editing", "D20_R2_percentage_editing"),
    # Then D4
    ("D4_R1_percentage_editing", "D4_R2_percentage_editing"),
    # Then D27
    ("D27_R1_percentage_editing", "D27_R2_percentage_editing"),
    # Then D34
    ("D34_R1_percentage_editing", "D34_R2_percentage_editing"),
    # Single-replicate timepoints
    ("D10obn_percentage_editing",),
    ("D34obn_percentage_editing",),
    ("D43obn_percentage_editing",),
]

MLH1_EFF_COLS = [
    ("PEmax_D20_percentage_editing",),
    ("PEmax_MLH1dn_D20_percentage_editing",),
    ("PEmax_obn_D20_percentage_editing",),
    ("PEmax_MLH1dn_obn_D20_percentage_editing",),
]


def load_excel_efficiency(filepath, sheet_name, eff_col_prefs):
    """Load efficiency data from Excel, keyed by 0-based data row index."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    # Row 1 = title, Row 2 = headers, Row 3+ = data
    headers = []
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        headers = [str(h).strip() if h else "" for h in row]

    # Map header names to column indices
    col_idx = {h: i for i, h in enumerate(headers) if h}

    # Build ordered list of efficiency column index groups
    eff_groups = []
    for group in eff_col_prefs:
        indices = []
        for col_name in group:
            if col_name in col_idx:
                indices.append(col_idx[col_name])
        if indices:
            eff_groups.append(indices)

    if not eff_groups:
        print(f"  WARNING: No efficiency columns found in {filepath}")
        wb.close()
        return {}

    print(f"  Efficiency column groups: {[[headers[i] for i in g] for g in eff_groups]}")

    # Read data rows (starting from row 3 = 0-indexed data row 0)
    row_efficiencies = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        row_vals = list(row)

        # Try each group in preference order
        best_eff = None
        for group_indices in eff_groups:
            vals = []
            for ci in group_indices:
                if ci < len(row_vals) and row_vals[ci] is not None:
                    try:
                        v = float(row_vals[ci])
                        vals.append(v)
                    except (ValueError, TypeError):
                        pass
            if vals:
                # Average the replicates in this group
                best_eff = sum(vals) / len(vals)
                break

        if best_eff is not None:
            row_efficiencies[row_idx] = round(best_eff, 4)

    wb.close()
    return row_efficiencies


def main():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Get all entries missing efficiency for paper 44
    cur.execute("""
        SELECT id, raw_source_text
        FROM pegrna_entries
        WHERE paper_id = 44 AND editing_efficiency IS NULL
    """)
    missing_entries = cur.fetchall()
    print(f"Total entries missing efficiency: {len(missing_entries)}")

    # Parse source file and row number from raw_source_text
    # Format: "Row 3 from mmc4.xlsx / Table_supp_SMARCB1-pegRNAData"
    pattern = re.compile(r"Row (\d+) from (mmc\d+\.xlsx) / (.+)")

    entries_by_file = {}
    for entry_id, raw_text in missing_entries:
        if not raw_text:
            continue
        m = pattern.match(raw_text)
        if m:
            row_num = int(m.group(1))
            filename = m.group(2)
            sheet = m.group(3)
            key = (filename, sheet)
            if key not in entries_by_file:
                entries_by_file[key] = []
            entries_by_file[key].append((entry_id, row_num))

    print(f"Entries by source file:")
    for (fname, sheet), entries in sorted(entries_by_file.items()):
        print(f"  {fname} / {sheet}: {len(entries)} entries")

    # Load efficiency data from each Excel file
    file_configs = {
        ("mmc4.xlsx", "Table_supp_SMARCB1-pegRNAData"): MMC4_EFF_COLS,
        ("mmc6.xlsx", "Supp_table_MLH1_x10_pegRNA_scor"): MLH1_EFF_COLS,
        ("mmc8.xlsx", "Supp_table_MLH1_non_coding_pegR"): MLH1_EFF_COLS,
    }

    total_updated = 0
    total_still_missing = 0

    for (fname, sheet), eff_prefs in file_configs.items():
        key = (fname, sheet)
        if key not in entries_by_file:
            print(f"\nSkipping {fname} / {sheet} - no missing entries")
            continue

        entries = entries_by_file[key]
        filepath = f"{RAW_DIR}/{fname}"
        print(f"\nProcessing {fname} / {sheet} ({len(entries)} missing)...")

        row_effs = load_excel_efficiency(filepath, sheet, eff_prefs)
        print(f"  Rows with efficiency data: {len(row_effs)}")

        updated = 0
        still_missing = 0
        for entry_id, row_num in entries:
            if row_num in row_effs:
                cur.execute(
                    "UPDATE pegrna_entries SET editing_efficiency = ? WHERE id = ?",
                    (row_effs[row_num], entry_id),
                )
                updated += 1
            else:
                still_missing += 1

        print(f"  Updated: {updated}, Still missing: {still_missing}")
        total_updated += updated
        total_still_missing += still_missing

    conn.commit()

    # Final stats
    cur.execute(
        "SELECT COUNT(*) FROM pegrna_entries WHERE paper_id = 44 AND editing_efficiency IS NOT NULL"
    )
    with_eff = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM pegrna_entries WHERE paper_id = 44")
    total = cur.fetchone()[0]
    print(f"\n=== Final Results ===")
    print(f"Total updated: {total_updated}")
    print(f"Still missing: {total_still_missing}")
    print(f"Paper 44 efficiency coverage: {with_eff}/{total} ({100*with_eff/total:.1f}%)")

    conn.close()


if __name__ == "__main__":
    main()
