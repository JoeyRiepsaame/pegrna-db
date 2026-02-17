#!/usr/bin/env python3
"""
Scan all downloaded supplementary xlsx files for columns that might contain
nicking sgRNA sequences.

Searches column headers (case-insensitive) for terms related to nicking guides,
PE3/PE5 systems, and secondary spacer/guide sequences.

Skips papers that are already being handled.
"""

import os
import re
import sys
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

RAW_PAPERS_DIR = Path("/Users/joeyriepsaame/pegrna-db/data/raw_papers")

# Papers to skip (already being handled)
SKIP_PAPERS = {
    "PMC11754097",
    "PMC12062269",
    "PMC12789627",
    "PMC10869272",
    "PMC12125189",
    "PMC11696010",
}

# Primary search terms - strong indicators of nicking guide columns
PRIMARY_TERMS = [
    r"nick",                # nick, nicking, nick_seq, nicking_guide, etc.
    r"n20",                 # N20 (nicking guide protospacer notation)
    r"pe3",                 # PE3 system (uses nicking sgRNA)
    r"pe5",                 # PE5 system (uses nicking sgRNA)
    r"nicking\s*guide",     # explicit "nicking guide"
    r"nick[_\s-]?seq",      # nick_seq, nick seq, nick-seq
    r"second\w*\s*g(uide|rna|sgrna)",  # second guide, secondary gRNA, etc.
    r"2nd\s*g(uide|rna|sgrna)",        # 2nd guide, 2nd gRNA
    r"ngRNA",               # nicking gRNA abbreviation
]

# Secondary search terms - columns that MIGHT be a separate guide/spacer
# (need manual review)
SECONDARY_TERMS = [
    r"spacer[\s_]?2",       # spacer_2, spacer 2
    r"guide[\s_]?2",        # guide_2, guide 2
    r"sgRNA[\s_]?2",        # sgRNA_2, sgRNA 2
    r"second\w*\s*spacer",  # secondary spacer, second spacer
    r"2nd\s*spacer",        # 2nd spacer
    r"auxiliary",           # auxiliary guide
    r"helper",              # helper guide
    r"secondary",           # secondary (in context of guides)
    r"sgrna.*nick",         # sgRNA nicking, sgRNA_nick
    r"cut\s*site\s*2",      # cut site 2
]

# Also broadly flag any column that mentions sgRNA/gRNA/guide/spacer
# but is NOT the main pegRNA spacer — these are candidates for review
BROAD_GUIDE_TERMS = [
    r"\bsgrna\b",
    r"\bgrna\b",
    r"\bguide\b",
    r"\bspacer\b",
    r"\bprotospacer\b",
]


def compile_patterns(terms: list[str]) -> list[re.Pattern]:
    return [re.compile(t, re.IGNORECASE) for t in terms]


PRIMARY_PATTERNS = compile_patterns(PRIMARY_TERMS)
SECONDARY_PATTERNS = compile_patterns(SECONDARY_TERMS)
BROAD_PATTERNS = compile_patterns(BROAD_GUIDE_TERMS)


def get_paper_id(filepath: Path) -> str | None:
    """Extract PMC ID from the file path."""
    for part in filepath.parts:
        if part.startswith("PMC"):
            return part
    return None


def scan_xlsx(filepath: Path) -> list[dict]:
    """
    Open an xlsx file, read every sheet's column headers, and return matches.
    """
    results = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        results.append({
            "file": str(filepath),
            "sheet": "N/A",
            "column": "N/A",
            "match_type": "ERROR",
            "detail": str(e),
        })
        return results

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            # Read the first row for headers
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(cell) if cell is not None else "" for cell in row]
                break

            if not headers:
                continue

            for col_idx, header in enumerate(headers):
                if not header or header == "":
                    continue

                header_clean = header.strip()

                # Check primary patterns
                for pat in PRIMARY_PATTERNS:
                    if pat.search(header_clean):
                        results.append({
                            "file": str(filepath),
                            "sheet": sheet_name,
                            "column": header_clean,
                            "col_index": col_idx,
                            "match_type": "PRIMARY",
                            "matched_pattern": pat.pattern,
                        })
                        break  # one primary match is enough per column

                # Check secondary patterns
                for pat in SECONDARY_PATTERNS:
                    if pat.search(header_clean):
                        results.append({
                            "file": str(filepath),
                            "sheet": sheet_name,
                            "column": header_clean,
                            "col_index": col_idx,
                            "match_type": "SECONDARY",
                            "matched_pattern": pat.pattern,
                        })
                        break

                # Check broad guide terms (only flag, lower priority)
                for pat in BROAD_PATTERNS:
                    if pat.search(header_clean):
                        # Avoid duplicating if already matched above
                        already = any(
                            r["column"] == header_clean
                            and r["sheet"] == sheet_name
                            and r["file"] == str(filepath)
                            for r in results
                        )
                        if not already:
                            results.append({
                                "file": str(filepath),
                                "sheet": sheet_name,
                                "column": header_clean,
                                "col_index": col_idx,
                                "match_type": "BROAD_GUIDE",
                                "matched_pattern": pat.pattern,
                            })
                        break

        except Exception as e:
            results.append({
                "file": str(filepath),
                "sheet": sheet_name,
                "column": "N/A",
                "match_type": "ERROR",
                "detail": str(e),
            })

    try:
        wb.close()
    except Exception:
        pass

    return results


def main():
    # Collect all xlsx files, skipping excluded papers
    xlsx_files = []
    for root, dirs, files in os.walk(RAW_PAPERS_DIR):
        for f in files:
            if f.endswith(".xlsx"):
                full_path = Path(root) / f
                paper_id = get_paper_id(full_path)
                if paper_id and paper_id in SKIP_PAPERS:
                    continue
                xlsx_files.append(full_path)

    xlsx_files.sort()

    print(f"{'='*80}")
    print(f"NICKING sgRNA COLUMN SCANNER")
    print(f"{'='*80}")
    print(f"Scanning {len(xlsx_files)} xlsx files...")
    print(f"Skipping papers: {', '.join(sorted(SKIP_PAPERS))}")
    print(f"{'='*80}\n")

    all_results = []
    files_scanned = 0
    files_with_errors = 0

    for xlsx_path in xlsx_files:
        paper_id = get_paper_id(xlsx_path) or "UNKNOWN"
        rel_path = xlsx_path.relative_to(RAW_PAPERS_DIR)
        print(f"  Scanning: {rel_path}")

        results = scan_xlsx(xlsx_path)
        all_results.extend(results)
        files_scanned += 1

        errors = [r for r in results if r["match_type"] == "ERROR"]
        if errors:
            files_with_errors += 1
            for err in errors:
                print(f"    ERROR: {err.get('detail', 'unknown error')}")

    # ---------------------------------------------------------------------------
    # Summary
    # ---------------------------------------------------------------------------
    primary_hits = [r for r in all_results if r["match_type"] == "PRIMARY"]
    secondary_hits = [r for r in all_results if r["match_type"] == "SECONDARY"]
    broad_hits = [r for r in all_results if r["match_type"] == "BROAD_GUIDE"]
    error_hits = [r for r in all_results if r["match_type"] == "ERROR"]

    print(f"\n{'='*80}")
    print(f"SCAN COMPLETE")
    print(f"{'='*80}")
    print(f"Files scanned: {files_scanned}")
    print(f"Files with errors: {files_with_errors}")
    print(f"PRIMARY matches (strong nicking indicators): {len(primary_hits)}")
    print(f"SECONDARY matches (possible nicking/2nd guide): {len(secondary_hits)}")
    print(f"BROAD matches (any guide/spacer/sgRNA column): {len(broad_hits)}")
    print()

    # ---- PRIMARY MATCHES ----
    if primary_hits:
        print(f"\n{'='*80}")
        print("PRIMARY MATCHES — Strong nicking sgRNA indicators")
        print(f"{'='*80}")
        current_file = None
        for r in primary_hits:
            rel = Path(r["file"]).relative_to(RAW_PAPERS_DIR)
            paper = get_paper_id(Path(r["file"])) or "?"
            if r["file"] != current_file:
                current_file = r["file"]
                print(f"\n  [{paper}] {rel}")
            print(f"    Sheet: {r['sheet']:30s}  Column: {r['column']}")
            print(f"      Matched: {r['matched_pattern']}")

    # ---- SECONDARY MATCHES ----
    if secondary_hits:
        print(f"\n{'='*80}")
        print("SECONDARY MATCHES — Possible nicking / 2nd guide columns")
        print(f"{'='*80}")
        current_file = None
        for r in secondary_hits:
            rel = Path(r["file"]).relative_to(RAW_PAPERS_DIR)
            paper = get_paper_id(Path(r["file"])) or "?"
            if r["file"] != current_file:
                current_file = r["file"]
                print(f"\n  [{paper}] {rel}")
            print(f"    Sheet: {r['sheet']:30s}  Column: {r['column']}")
            print(f"      Matched: {r['matched_pattern']}")

    # ---- BROAD MATCHES (condensed) ----
    if broad_hits:
        print(f"\n{'='*80}")
        print("BROAD MATCHES — All guide/spacer/sgRNA columns (for review)")
        print(f"{'='*80}")
        current_file = None
        for r in broad_hits:
            rel = Path(r["file"]).relative_to(RAW_PAPERS_DIR)
            paper = get_paper_id(Path(r["file"])) or "?"
            if r["file"] != current_file:
                current_file = r["file"]
                print(f"\n  [{paper}] {rel}")
            print(f"    Sheet: {r['sheet']:30s}  Column: {r['column']}")

    # ---- ERRORS ----
    if error_hits:
        print(f"\n{'='*80}")
        print("ERRORS")
        print(f"{'='*80}")
        for r in error_hits:
            rel = Path(r["file"]).relative_to(RAW_PAPERS_DIR)
            print(f"  {rel}: {r.get('detail', 'unknown')}")

    # ---- ALL COLUMN HEADERS (for completeness) ----
    print(f"\n{'='*80}")
    print("FULL COLUMN HEADER LISTING (all sheets, all files)")
    print(f"{'='*80}")
    for xlsx_path in xlsx_files:
        rel = xlsx_path.relative_to(RAW_PAPERS_DIR)
        paper_id = get_paper_id(xlsx_path) or "?"
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = []
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    headers = [str(cell) if cell is not None else "" for cell in row]
                    break
                if headers:
                    print(f"\n  [{paper_id}] {rel} -> Sheet: {sheet_name}")
                    for i, h in enumerate(headers):
                        if h:
                            print(f"    [{i:3d}] {h}")
            wb.close()
        except Exception as e:
            print(f"\n  [{paper_id}] {rel} -> ERROR: {e}")

    print(f"\n{'='*80}")
    print("DONE")
    print(f"{'='*80}")


if __name__ == "__main__":
    main()
