#!/usr/bin/env python3
"""
Fix extension_sequence for + strand PAM targets in PMC12008803 (paper_id=44).

The PBS and RTT columns in the Excel are stored in opposite orientation for
+ strand vs - strand PAM targets:
  - strand: extension = RTT + PBS (correct, matches PBS_RTT_5to3)
  + strand: extension = reverse(PBS + RTT) (PBS_RTT_5to3 column)

Our DB computed extension_sequence = rtt + pbs for ALL entries, which is wrong
for the 9,148 entries targeting the + strand.

Fix: use the actual PBS_RTT_5to3 value from Excel for + strand entries.
"""

import re
import sqlite3
import openpyxl

DB_PATH = "data/pegrna_database.db"
RAW_DIR = "data/raw_papers/PMC12008803/PMC12008803"

SCAFFOLD = "GTTTAAGAGCTATGCTGGAAACAGCATAGCAAGTTTAAATAAGGCTAGTCCGTTATCAACTTGAAAAAGTGGCACCGAGTCGGTGC"

FILES = {
    "mmc2.xlsx": "_SuppTable_ScaffoldComparison",
    "mmc4.xlsx": "Table_supp_SMARCB1-pegRNAData",
    "mmc6.xlsx": "Supp_table_MLH1_x10_pegRNA_scor",
    "mmc8.xlsx": "Supp_table_MLH1_non_coding_pegR",
}


def load_excel_extensions(filepath, sheet_name):
    """Load PBS_RTT_5to3 and PAM strand from Excel, keyed by 0-based row index."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    headers = []
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        headers = [str(h).strip() if h else "" for h in row]
    ci = {h: i for i, h in enumerate(headers)}

    data = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        strand = row[ci["PAM strand"]] if "PAM strand" in ci else None
        pbs_rtt = str(row[ci["PBS_RTT_5to3"]]) if row[ci.get("PBS_RTT_5to3", -1)] else None

        if strand == "+" and pbs_rtt:
            data[row_idx] = pbs_rtt.strip().upper()

    wb.close()
    return data


def main():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Get all entries for paper 44
    cur.execute("""
        SELECT id, raw_source_text, spacer_sequence, extension_sequence
        FROM pegrna_entries WHERE paper_id = 44
    """)
    all_entries = cur.fetchall()
    print(f"Total entries for paper 44: {len(all_entries)}")

    # Parse source file and row number
    pattern = re.compile(r"Row (\d+) from (mmc\d+\.xlsx) / (.+)")
    entries_by_file = {}
    for entry_id, raw_text, spacer, ext in all_entries:
        if not raw_text:
            continue
        m = pattern.match(raw_text)
        if m:
            row_num = int(m.group(1))
            filename = m.group(2)
            sheet = m.group(3)
            key = (filename, sheet)
            if key not in entries_by_file:
                entries_by_file[key] = []
            entries_by_file[key].append((entry_id, row_num, spacer, ext))

    total_fixed_ext = 0
    total_fixed_full = 0

    for fname, sheet in FILES.items():
        key = (fname, sheet)
        if key not in entries_by_file:
            print(f"\nSkipping {fname} - no entries found")
            continue

        entries = entries_by_file[key]
        filepath = f"{RAW_DIR}/{fname}"
        print(f"\nProcessing {fname} / {sheet}...")

        plus_strand_exts = load_excel_extensions(filepath, sheet)
        print(f"  + strand rows with PBS_RTT_5to3: {len(plus_strand_exts)}")

        fixed_ext = 0
        fixed_full = 0
        for entry_id, row_num, spacer, old_ext in entries:
            if row_num not in plus_strand_exts:
                continue  # - strand or missing data, skip

            correct_ext = plus_strand_exts[row_num]

            if old_ext and old_ext.upper() == correct_ext:
                continue  # Already correct

            # Update extension_sequence
            cur.execute(
                "UPDATE pegrna_entries SET extension_sequence = ? WHERE id = ?",
                (correct_ext, entry_id),
            )
            fixed_ext += 1

            # Also reconstruct full_sequence if we have spacer
            if spacer:
                full_seq = spacer + SCAFFOLD + correct_ext
                cur.execute(
                    "UPDATE pegrna_entries SET full_sequence = ? WHERE id = ?",
                    (full_seq, entry_id),
                )
                fixed_full += 1

        print(f"  Fixed extension_sequence: {fixed_ext}")
        print(f"  Fixed full_sequence: {fixed_full}")
        total_fixed_ext += fixed_ext
        total_fixed_full += fixed_full

    conn.commit()

    # Verify: count entries where extension != rtt+pbs (should be the + strand ones)
    cur.execute("""
        SELECT COUNT(*) FROM pegrna_entries
        WHERE paper_id = 44
        AND rtt_sequence IS NOT NULL AND pbs_sequence IS NOT NULL
        AND extension_sequence != (rtt_sequence || pbs_sequence)
    """)
    diff_count = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(*) FROM pegrna_entries
        WHERE paper_id = 44 AND full_sequence IS NOT NULL
    """)
    full_count = cur.fetchone()[0]

    print(f"\n=== Results ===")
    print(f"Extension sequences fixed: {total_fixed_ext}")
    print(f"Full sequences reconstructed: {total_fixed_full}")
    print(f"Entries where extension != rtt+pbs (+ strand): {diff_count}")
    print(f"Entries with full_sequence: {full_count}")

    conn.close()


if __name__ == "__main__":
    main()
